from typing import Any, Dict, List

from openpyxl import load_workbook

from .base import BaseLoader
from ._text import _strip_cell


class XlsxLoader(BaseLoader):
    name = "xlsx"

    def load(self, path) -> List[Dict[str, Any]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows = list(_iter_rows(ws))
        finally:
            wb.close()
        if not rows:
            return []
        header = [_strip_cell(cell) for cell in rows[0]]
        result = []
        for row in rows[1:]:
            if not any(str(cell).strip() for cell in row):
                continue
            padded = list(row) + [""] * (len(header) - len(row))
            record = {}
            for name, value in zip(header, padded):
                if name:
                    record[name] = _strip_cell(value)
            if any(record.values()):
                result.append(record)
        return result


def _iter_rows(ws):
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            yield row
